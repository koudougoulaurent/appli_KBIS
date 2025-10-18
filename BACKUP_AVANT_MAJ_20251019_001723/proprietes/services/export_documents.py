"""
Service d'export des documents pour le module d'archivage
"""
import io
import os
from datetime import datetime
from typing import List, Dict, Any, Optional
from django.http import HttpResponse
from django.db.models import QuerySet
from django.utils import timezone
from django.conf import settings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from PIL import Image
import tempfile

from ..models import Document


class DocumentExportService:
    """Service pour l'export des documents en différents formats."""
    
    def __init__(self):
        self.styles = self._get_excel_styles()
    
    def _get_excel_styles(self) -> Dict[str, Any]:
        """Définit les styles pour Excel."""
        return {
            'header': Font(name='Arial', size=12, bold=True, color='FFFFFF'),
            'header_fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'subheader': Font(name='Arial', size=11, bold=True, color='2F5597'),
            'normal': Font(name='Arial', size=10),
            'date': Font(name='Arial', size=10, color='666666'),
            'center': Alignment(horizontal='center', vertical='center'),
            'left': Alignment(horizontal='left', vertical='center'),
            'right': Alignment(horizontal='right', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    def export_to_excel(self, documents: QuerySet, filename: Optional[str] = None) -> HttpResponse:
        """Exporte les documents vers un fichier Excel."""
        if not filename:
            filename = f"documents_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Créer un workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Documents"
        
        # En-têtes
        headers = [
            'ID', 'Nom', 'Type', 'Description', 'Propriété', 'Bailleur', 'Locataire',
            'Statut', 'Date Création', 'Date Modification', 'Date Expiration',
            'Créé par', 'Tags', 'Confidentiel', 'Taille (KB)'
        ]
        
        # Appliquer les styles aux en-têtes
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.styles['header']
            cell.fill = self.styles['header_fill']
            cell.alignment = self.styles['center']
            cell.border = self.styles['border']
        
        # Données des documents
        row = 2
        for doc in documents:
            ws.cell(row=row, column=1, value=doc.id)
            ws.cell(row=row, column=2, value=doc.nom)
            ws.cell(row=row, column=3, value=doc.get_type_document_display())
            ws.cell(row=row, column=4, value=doc.description or '')
            ws.cell(row=row, column=5, value=str(doc.propriete) if doc.propriete else '')
            ws.cell(row=row, column=6, value=str(doc.bailleur) if doc.bailleur else '')
            ws.cell(row=row, column=7, value=str(doc.locataire) if doc.locataire else '')
            ws.cell(row=row, column=8, value=doc.get_statut_display())
            ws.cell(row=row, column=9, value=doc.date_creation.strftime('%d/%m/%Y %H:%M'))
            ws.cell(row=row, column=10, value=doc.date_modification.strftime('%d/%m/%Y %H:%M'))
            ws.cell(row=row, column=11, value=doc.date_expiration.strftime('%d/%m/%Y') if doc.date_expiration else '')
            ws.cell(row=row, column=12, value=str(doc.cree_par) if doc.cree_par else '')
            ws.cell(row=row, column=13, value=doc.tags or '')
            ws.cell(row=row, column=14, value='Oui' if doc.confidentiel else 'Non')
            ws.cell(row=row, column=15, value=round(doc.taille_fichier / 1024, 2) if doc.taille_fichier else 0)
            
            # Appliquer les styles aux données
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = self.styles['normal']
                cell.alignment = self.styles['left']
                cell.border = self.styles['border']
            
            row += 1
        
        # Ajuster la largeur des colonnes
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            if col == 1:  # ID
                ws.column_dimensions[column_letter].width = 8
            elif col == 2:  # Nom
                ws.column_dimensions[column_letter].width = 30
            elif col == 4:  # Description
                ws.column_dimensions[column_letter].width = 40
            elif col in [5, 6, 7]:  # Propriété, Bailleur, Locataire
                ws.column_dimensions[column_letter].width = 25
            elif col in [9, 10, 11]:  # Dates
                ws.column_dimensions[column_letter].width = 18
            else:
                ws.column_dimensions[column_letter].width = 15
        
        # Créer la réponse HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Sauvegarder le workbook
        with io.BytesIO() as output:
            wb.save(output)
            response.write(output.getvalue())
        
        return response
    
    def export_to_pdf(self, documents: QuerySet, filename: Optional[str] = None) -> HttpResponse:
        """Exporte les documents vers un fichier PDF."""
        if not filename:
            filename = f"documents_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        # Créer la réponse HTTP
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Créer le document PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#2F5597')
        )
        
        header_style = ParagraphStyle(
            'CustomHeader',
            parent=styles['Heading2'],
            fontSize=12,
            spaceAfter=12,
            textColor=colors.HexColor('#366092')
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=6
        )
        
        # Contenu du PDF
        story = []
        
        # Titre
        story.append(Paragraph("Rapport d'Archivage des Documents", title_style))
        story.append(Spacer(1, 12))
        
        # Informations générales
        story.append(Paragraph(f"<b>Date d'export:</b> {datetime.now().strftime('%d/%m/%Y à %H:%M')}", normal_style))
        story.append(Paragraph(f"<b>Nombre de documents:</b> {documents.count()}", normal_style))
        story.append(Spacer(1, 20))
        
        # Tableau des documents
        if documents.exists():
            # En-têtes du tableau
            table_data = [['Nom', 'Type', 'Propriété', 'Statut', 'Date Création', 'Taille']]
            
            # Données des documents
            for doc in documents:
                table_data.append([
                    doc.nom[:30] + '...' if len(doc.nom) > 30 else doc.nom,
                    doc.get_type_document_display(),
                    str(doc.propriete)[:20] + '...' if doc.propriete and len(str(doc.propriete)) > 20 else str(doc.propriete) if doc.propriete else '',
                    doc.get_statut_display(),
                    doc.date_creation.strftime('%d/%m/%Y'),
                    f"{round(doc.taille_fichier / 1024, 1)} KB" if doc.taille_fichier else "0 KB"
                ])
            
            # Créer le tableau
            table = Table(table_data, colWidths=[2*inch, 1*inch, 1.5*inch, 0.8*inch, 1*inch, 0.8*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))
            
            story.append(table)
        else:
            story.append(Paragraph("Aucun document trouvé.", normal_style))
        
        # Construire le PDF
        doc.build(story)
        
        # Récupérer le contenu
        pdf_content = buffer.getvalue()
        buffer.close()
        
        response.write(pdf_content)
        return response
    
    def export_to_csv(self, documents: QuerySet, filename: Optional[str] = None) -> HttpResponse:
        """Exporte les documents vers un fichier CSV."""
        import csv
        
        if not filename:
            filename = f"documents_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Ajouter BOM pour Excel
        response.write('\ufeff')
        
        writer = csv.writer(response, delimiter=';')
        
        # En-têtes
        writer.writerow([
            'ID', 'Nom', 'Type', 'Description', 'Propriété', 'Bailleur', 'Locataire',
            'Statut', 'Date Création', 'Date Modification', 'Date Expiration',
            'Créé par', 'Tags', 'Confidentiel', 'Taille (KB)'
        ])
        
        # Données
        for doc in documents:
            writer.writerow([
                doc.id,
                doc.nom,
                doc.get_type_document_display(),
                doc.description or '',
                str(doc.propriete) if doc.propriete else '',
                str(doc.bailleur) if doc.bailleur else '',
                str(doc.locataire) if doc.locataire else '',
                doc.get_statut_display(),
                doc.date_creation.strftime('%d/%m/%Y %H:%M'),
                doc.date_modification.strftime('%d/%m/%Y %H:%M'),
                doc.date_expiration.strftime('%d/%m/%Y') if doc.date_expiration else '',
                str(doc.cree_par) if doc.cree_par else '',
                doc.tags or '',
                'Oui' if doc.confidentiel else 'Non',
                round(doc.taille_fichier / 1024, 2) if doc.taille_fichier else 0
            ])
        
        return response
    
    def get_export_statistics(self, documents: QuerySet) -> Dict[str, Any]:
        """Calcule les statistiques pour l'export."""
        # Convertir en liste pour éviter les problèmes de slice
        docs_list = list(documents)
        total_docs = len(docs_list)
        
        # Statistiques par type
        type_stats = {}
        for doc_type, _ in Document.TYPE_DOCUMENT_CHOICES:
            count = sum(1 for doc in docs_list if doc.type_document == doc_type)
            if count > 0:
                type_stats[doc_type] = count
        
        # Statistiques par statut
        statut_stats = {}
        for statut, _ in Document.STATUT_CHOICES:
            count = sum(1 for doc in docs_list if doc.statut == statut)
            if count > 0:
                statut_stats[statut] = count
        
        # Taille totale
        total_size = sum(doc.taille_fichier or 0 for doc in docs_list)
        
        return {
            'total_documents': total_docs,
            'type_statistics': type_stats,
            'statut_statistics': statut_stats,
            'total_size_bytes': total_size,
            'total_size_mb': round(total_size / (1024 * 1024), 2),
            'export_date': timezone.now()
        }
