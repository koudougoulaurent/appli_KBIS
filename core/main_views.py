from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods, require_POST
from django.views.decorators.csrf import requires_csrf_token
from django.conf import settings
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum
from django.db.models.functions import ExtractHour, ExtractWeekDay
from django.utils import timezone
from datetime import datetime, timedelta
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import user_passes_test
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
from django.template.loader import render_to_string
import json
import os

from .models import ConfigurationEntreprise, TemplateRecu, Devise, AuditLog
from paiements.models import Paiement
from proprietes.models import Propriete, Locataire, Bailleur
from contrats.models import Contrat
from utilisateurs.models import Utilisateur
from .forms import ConfigurationEntrepriseForm
from .utils import convertir_montant, check_group_permissions
from django.contrib.contenttypes.models import ContentType
from .optimizations import (
    performance_monitor, 
    query_optimizer,
    get_cached_data,
    set_cached_data
)

Utilisateur = get_user_model()

def is_admin(user):
    """Vérifie si l'utilisateur appartient au groupe PRIVILEGE."""
    return user.is_authenticated and hasattr(user, 'groupe_travail') and user.groupe_travail.nom.upper() == 'PRIVILEGE'

@login_required
@performance_monitor
def home(request):
    """Page d'accueil du système - OPTIMISÉE."""
    # Utiliser le cache pour les statistiques
    user_id = request.user.id if request.user.is_authenticated else None
    cache_key = f"home_stats_{user_id}"
    cached_stats = get_cached_data(cache_key, None, 300)  # 5 minutes
    
    if not cached_stats:
        # Calculer les statistiques de base
        stats = {}
        
        # Récupérer les données récentes
        recent_data = {'paiements_recents': []}
        
        cached_stats = {
            'stats': stats,
            'paiements_recents': recent_data['paiements_recents'],
        }
        
        # Mettre en cache
        from django.core.cache import cache
        cache.set(cache_key, cached_stats, 300)
    
    # Préparer le contexte du template
    context = cached_stats
    
    return render(request, 'core/home.html', context)

@login_required
@performance_monitor
def dashboard(request):
    """Tableau de bord principal unifié - SÉCURISÉ (sans informations confidentielles)."""
    # Utiliser le cache pour les statistiques du dashboard
    user_id = request.user.id
    cache_key = f"dashboard_stats_{user_id}"
    cached_stats = get_cached_data(cache_key, None, 600)  # 10 minutes
    
    # Initialiser avec des valeurs par défaut
    if not cached_stats:
        # Calculer les statistiques de base pour tous les modules (UNIQUEMENT non confidentielles)
        stats = {}
        
        # Statistiques des propriétés (non confidentielles)
        total_proprietes = Propriete.objects.filter(is_deleted=False).count()
        proprietes_louees = Propriete.objects.filter(is_deleted=False, disponible=False).count()
        proprietes_disponibles = Propriete.objects.filter(is_deleted=False, disponible=True).count()
        proprietes_en_construction = 0  # Pas de champ construction dans le modèle actuel
        
        # Statistiques des bailleurs et locataires (non confidentielles)
        total_bailleurs = Bailleur.objects.filter(is_deleted=False).count()
        bailleurs_actifs = Bailleur.objects.filter(is_deleted=False, actif=True).count()
        total_locataires = Locataire.objects.filter(is_deleted=False).count()
        locataires_actifs = Locataire.objects.filter(is_deleted=False, statut='actif').count()
        
        # Statistiques des contrats (non confidentielles)
        total_contrats = Contrat.objects.filter(is_deleted=False).count()
        contrats_actifs = Contrat.objects.filter(is_deleted=False, est_actif=True).count()
        
        # Statistiques des paiements (UNIQUEMENT le nombre, PAS les montants)
        total_paiements = Paiement.objects.filter(is_deleted=False).count()
        # Supprimer le comptage mensuel des paiements pour éviter de révéler l'activité financière
        
        # Tendances générales (non confidentielles)
        tendances = {
            'activite_generale': 'stable',  # Valeur générique
        }
        
        # NE PAS inclure d'informations sur les devises ou montants
        # Supprimer toutes les références financières
        
        cached_stats = {
            'stats': stats,
            'tendances': tendances,
            # Statistiques des propriétés (non confidentielles)
            'total_proprietes': total_proprietes,
            'proprietes_louees': proprietes_louees,
            'proprietes_disponibles': proprietes_disponibles,
            'proprietes_en_construction': proprietes_en_construction,
            # Statistiques des bailleurs et locataires (non confidentielles)
            'total_bailleurs': total_bailleurs,
            'bailleurs_actifs': bailleurs_actifs,
            'total_locataires': total_locataires,
            'locataires_actifs': locataires_actifs,
            # Statistiques des contrats (non confidentielles)
            'total_contrats': total_contrats,
            'contrats_actifs': contrats_actifs,
            # Statistiques des paiements (UNIQUEMENT le nombre total)
            'total_paiements': total_paiements,
            # SUPPRIMER: paiements_mois, devise_base, devise_active
        }
        
        # Mettre en cache
        from django.core.cache import cache
        cache.set(cache_key, cached_stats, 600)
    
    # Préparer le contexte du template
    # S'assurer que cached_stats a toujours une valeur
    if cached_stats is None:
        cached_stats = {}
    
    context = cached_stats
    
    return render(request, 'core/dashboard_unified.html', context)

@login_required
def intelligent_search(request):
    """Recherche intelligente."""
    query = request.GET.get('q', '')
    results = []
    
    if query:
        # Recherche dans les paiements (SANS les montants pour la confidentialité)
        paiements = Paiement.objects.filter(
            Q(contrat__numero_contrat__icontains=query) |
            Q(contrat__locataire__nom__icontains=query) |
            Q(contrat__locataire__prenom__icontains=query)
            # SUPPRIMER: Q(montant__icontains=query) - Information confidentielle
        ).select_related('contrat__locataire')[:10]
        
        # Recherche dans les propriétés
        proprietes = Propriete.objects.filter(
            Q(adresse__icontains=query) |
            Q(bailleur__nom__icontains=query)
        ).select_related('bailleur')[:10]
        
        results = {
            'paiements': paiements,
            'proprietes': proprietes,
        }
    
    context = {
        'query': query,
        'results': results,
    }
    
    return render(request, 'core/intelligent_search.html', context)

# Configuration de l'entreprise
@login_required
def configuration_entreprise(request):
    """
    Vue pour configurer les informations de l'entreprise
    """
    # Vérification des permissions : Tous les utilisateurs connectés peuvent accéder
    # mais seuls PRIVILEGE peuvent modifier
    from core.utils import check_group_permissions
    can_modify = check_group_permissions(request.user, ['PRIVILEGE'], 'modify')['allowed']
    
    # Récupérer ou créer la configuration active
    config = ConfigurationEntreprise.get_configuration_active()
    if not config:
        config = ConfigurationEntreprise.objects.create(
            nom_entreprise="GESTIMMOB",
            adresse="123 Rue de la Paix",
            code_postal="75001",
            ville="Paris",
            pays="France",
            telephone="01 23 45 67 89",
            email="contact@gestimmob.fr",
            siret="123 456 789 00012",
            numero_licence="123456789"
        )
    
    if request.method == 'POST' and can_modify:
        form = ConfigurationEntrepriseForm(request.POST, instance=config)
        if form.is_valid():
            form.save()
            messages.success(request, 'Configuration de l\'entreprise mise à jour avec succès.')
            return redirect('core:configuration_entreprise')
        else:
            messages.error(request, 'Erreur dans le formulaire. Veuillez vérifier les informations.')
    elif request.method == 'POST' and not can_modify:
        messages.warning(request, 'Vous n\'avez pas les permissions pour modifier la configuration. Contactez un administrateur.')
        return redirect('core:configuration_entreprise')
    else:
        form = ConfigurationEntrepriseForm(instance=config)
    
    context = {
        'form': form,
        'config': config,
        'can_modify': can_modify,
    }
    
    return render(request, 'core/configuration_entreprise.html', context)

@login_required
def gestion_templates(request):
    """Gestion des templates de reçus."""
    templates = TemplateRecu.get_templates_actifs()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'ajouter':
            # Vérification des permissions : PRIVILEGE et ADMINISTRATION peuvent ajouter
            from core.utils import check_group_permissions
            permissions = check_group_permissions(request.user, ['PRIVILEGE'], 'add')
            if not permissions['allowed']:
                messages.error(request, permissions['message'])
                return redirect('core:gestion_templates')
            
            nom = request.POST.get('nom')
            description = request.POST.get('description')
            fichier = request.FILES.get('fichier_template')
            
            if nom and fichier:
                template = TemplateRecu.objects.create(
                    nom=nom,
                    description=description,
                    fichier_template=fichier
                )
                messages.success(request, f'Template "{nom}" créé avec succès !')
                return redirect('core:gestion_templates')
        
        elif action == 'modifier':
            # Vérification des permissions : Seul PRIVILEGE peut modifier
            from core.utils import check_group_permissions
            permissions = check_group_permissions(request.user, ['PRIVILEGE'], 'modify')
            if not permissions['allowed']:
                messages.error(request, permissions['message'])
                return redirect('core:gestion_templates')
            
            template_id = request.POST.get('template_id')
            template = get_object_or_404(TemplateRecu, id=template_id)
            
            template.nom = request.POST.get('nom', template.nom)
            template.description = request.POST.get('description', template.description)
            template.couleur_principale = request.POST.get('couleur_principale', template.couleur_principale)
            template.couleur_secondaire = request.POST.get('couleur_secondaire', template.couleur_secondaire)
            template.police_principale = request.POST.get('police_principale', template.police_principale)
            
            template.afficher_logo = 'afficher_logo' in request.POST
            template.afficher_siret = 'afficher_siret' in request.POST
            template.afficher_tva = 'afficher_tva' in request.POST
            template.afficher_iban = 'afficher_iban' in request.POST
            
            if 'fichier_template' in request.FILES:
                template.fichier_template = request.FILES['fichier_template']
            
            template.save()
            messages.success(request, f'Template "{template.nom}" modifié avec succès !')
            return redirect('core:gestion_templates')
        
        elif action == 'supprimer':
            # Vérification des permissions : Seul PRIVILEGE peut supprimer
            from core.utils import check_group_permissions
            permissions = check_group_permissions(request.user, ['PRIVILEGE'], 'delete')
            if not permissions['allowed']:
                messages.error(request, permissions['message'])
                return redirect('core:gestion_templates')
            
            template_id = request.POST.get('template_id')
            template = get_object_or_404(TemplateRecu, id=template_id)
            nom = template.nom
            old_data = {f.name: getattr(template, f.name) for f in template._meta.fields}
            template.is_deleted = True
            template.deleted_at = timezone.now()
            template.deleted_by = request.user
            template.save()
            AuditLog.objects.create(
                content_type=ContentType.objects.get_for_model(TemplateRecu),
                object_id=template.pk,
                action='delete',
                details={
                    'old_data': old_data,
                    'new_data': None
                },
                object_repr=str(template),
                user=request.user,
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            messages.success(request, f'Template "{nom}" supprimé avec succès (suppression logique) !')
            return redirect('core:gestion_templates')
        
        elif action == 'defaut':
            template_id = request.POST.get('template_id')
            template = get_object_or_404(TemplateRecu, id=template_id)
            template.par_defaut = True
            template.save()
            messages.success(request, f'Template "{template.nom}" défini par défaut !')
            return redirect('core:gestion_templates')
    
    context = {
        'templates': templates,
    }
    
    return render(request, 'core/gestion_templates.html', context)

@login_required
def apercu_template(request, template_id):
    """Aperçu d'un template de reçu."""
    # Vérification des permissions : PRIVILEGE, ADMINISTRATION peuvent visualiser
    from core.utils import check_group_permissions
    permissions = check_group_permissions(request.user, ['PRIVILEGE', 'ADMINISTRATION'], 'view')
    if not permissions['allowed']:
        messages.error(request, permissions['message'])
        return redirect('core:gestion_templates')
    
    template = get_object_or_404(TemplateRecu, id=template_id)
    config = ConfigurationEntreprise.get_configuration_active()
    
    # Données d'exemple pour l'aperçu
    paiement_exemple = Paiement.objects.select_related(
        'contrat__locataire',
        'contrat__propriete__bailleur'
    ).first()
    
    # Plus de génération automatique de reçus
    recu_exemple = None
    
    context = {
        'template': template,
        'config': config,
        'recu_exemple': recu_exemple,
        'paiement_exemple': paiement_exemple,
    }
    
    return render(request, 'core/apercu_template.html', context)

@login_required
def test_template(request, template_id):
    """Test d'un template avec génération PDF."""
    # Vérification des permissions : PRIVILEGE, ADMINISTRATION peuvent tester
    from core.utils import check_group_permissions
    permissions = check_group_permissions(request.user, ['PRIVILEGE', 'ADMINISTRATION'], 'view')
    if not permissions['allowed']:
        messages.error(request, permissions['message'])
        return redirect('core:gestion_templates')
    
    template = get_object_or_404(TemplateRecu, id=template_id)
    config = ConfigurationEntreprise.get_configuration_active()
    
    # Créer un paiement de test
    paiement_test = Paiement.objects.select_related(
        'contrat__locataire',
        'contrat__propriete__bailleur'
    ).first()
    
    # Fonctionnalité de test temporairement désactivée après refactoring
    messages.info(request, 'Fonctionnalité de test temporairement indisponible après refactoring du système de paiements.')
    return redirect('core:gestion_templates')
    
    messages.error(request, 'Aucun paiement disponible pour le test.')
    return redirect('core:gestion_templates')

# API pour la configuration
@login_required
@require_http_methods(["GET"])
def api_configuration(request):
    """API pour récupérer la configuration de l'entreprise."""
    config = ConfigurationEntreprise.get_configuration_active()
    
    data = {
        'nom_entreprise': config.nom_entreprise,
        'slogan': config.slogan,
        'adresse': config.adresse,
        'code_postal': config.code_postal,
        'ville': config.ville,
        'pays': config.pays,
        'telephone': config.telephone,
        'email': config.email,
        'site_web': config.site_web,
        'siret': config.siret,
        'numero_licence': config.numero_licence,
        'capital_social': config.capital_social,
        'forme_juridique': config.forme_juridique,
        'logo_url': config.logo_url,
        'couleur_principale': config.couleur_principale,
        'couleur_secondaire': config.couleur_secondaire,
        'iban': config.iban,
        'bic': config.bic,
        'banque': config.banque,
    }
    
    return JsonResponse(data)

@login_required
@require_http_methods(["POST"])
def api_sauvegarder_configuration(request):
    """API pour sauvegarder la configuration."""
    if not is_admin(request.user):
        return JsonResponse({'error': 'Permission refusée'}, status=403)
    
    try:
        data = json.loads(request.body)
        config = ConfigurationEntreprise.get_configuration_active()
        
        # Mise à jour des champs
        for field, value in data.items():
            if hasattr(config, field):
                setattr(config, field, value)
        
        config.full_clean()
        config.save()
        
        return JsonResponse({'success': True, 'message': 'Configuration sauvegardée'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@require_POST
def changer_devise(request):
    code = request.POST.get('devise')
    if code and Devise.objects.filter(code=code, actif=True).exists():
        request.session['devise_active'] = code
    return redirect(request.META.get('HTTP_REFERER', '/'))

def liste_devises(request):
    """Affiche la liste des devises disponibles dans le système."""
    devises = Devise.objects.filter(actif=True)
    context = {
        'devises': devises,
        'devise_active': getattr(request, 'devise_active', None),
    }
    return render(request, 'core/liste_devises.html', context)


@login_required
def api_interface(request):
    """
    Interface de test pour l'API REST
    """
    return render(request, 'api/interface.html')


@requires_csrf_token
def csrf_failure(request, reason=""):
    """
    Vue personnalisée pour gérer les erreurs CSRF
    """
    context = {
        'reason': reason,
        'DEBUG': settings.DEBUG,
    }
    return render(request, '403_csrf.html', context, status=403)

@login_required
def rapports_audit(request):
    """
    Vue pour afficher les rapports d'audit du système avec filtres avancés
    """
    # Vérification des permissions : PRIVILEGE, CONTROLES peuvent accéder aux rapports d'audit
    from core.utils import check_group_permissions
    permissions = check_group_permissions(request.user, ['PRIVILEGE', 'CONTROLES'], 'view')
    if not permissions['allowed']:
        messages.error(request, permissions['message'])
        return redirect('core:dashboard')
    
    # Récupération des paramètres de filtrage
    search_query = request.GET.get('search', '')
    action_type = request.GET.get('action_type', '')
    user_filter = request.GET.get('user', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    export_format = request.GET.get('export', '')
    page_size = int(request.GET.get('page_size', 20))
    
    # Construction du queryset de base
    queryset = AuditLog.objects.select_related('user', 'content_type')
    
    # Application des filtres
    if search_query:
        queryset = queryset.filter(
            Q(user__username__icontains=search_query) |
            Q(object_repr__icontains=search_query) |
            Q(details__icontains=search_query) |
            Q(ip_address__icontains=search_query)
        )
    
    if action_type:
        queryset = queryset.filter(action=action_type)
    
    if user_filter:
        queryset = queryset.filter(user__username=user_filter)
    
    if date_from:
        queryset = queryset.filter(timestamp__date__gte=date_from)
    
    if date_to:
        queryset = queryset.filter(timestamp__date__lte=date_to)
    
    # Tri par défaut
    queryset = queryset.order_by('-timestamp')
    
    # Export des données si demandé
    if export_format:
        return export_audit_data(queryset, export_format)
    
    # Pagination
    paginator = Paginator(queryset, page_size)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistiques d'audit
    stats_audit = {
        'total_actions': AuditLog.objects.count(),
        'actions_aujourd_hui': AuditLog.objects.filter(
            timestamp__date=timezone.now().date()
        ).count(),
        'actions_semaine': AuditLog.objects.filter(
            timestamp__gte=timezone.now() - timedelta(days=7)
        ).count(),
        'actions_mois': AuditLog.objects.filter(
            timestamp__gte=timezone.now() - timedelta(days=30)
        ).count(),
        'actions_an': AuditLog.objects.filter(
            timestamp__gte=timezone.now() - timedelta(days=365)
        ).count(),
    }
    
    # Actions par type avec pourcentages
    actions_par_type = AuditLog.objects.values('action').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # Calcul des pourcentages
    total_actions = stats_audit['total_actions']
    for action in actions_par_type:
        action['percentage'] = round((action['count'] / total_actions) * 100, 1) if total_actions > 0 else 0
    
    # Actions par utilisateur
    actions_par_utilisateur = AuditLog.objects.values('user__username').annotate(
        count=Count('id')
    ).filter(user__username__isnull=False).order_by('-count')[:10]
    
    # Actions par heure de la journée (pour graphique)
    actions_par_heure = AuditLog.objects.annotate(
        heure=ExtractHour('timestamp')
    ).values('heure').annotate(
        count=Count('id')
    ).order_by('heure')
    
    # Actions par jour de la semaine
    actions_par_jour = AuditLog.objects.annotate(
        jour=ExtractWeekDay('timestamp')
    ).values('jour').annotate(
        count=Count('id')
    ).order_by('jour')
    
    # Dernières actions critiques (suppressions, modifications importantes)
    actions_critiques = AuditLog.objects.filter(
        action__in=['delete', 'update']
    ).select_related('user', 'content_type').order_by('-timestamp')[:5]
    
    # Utilisateurs les plus actifs aujourd'hui
    utilisateurs_actifs_aujourd_hui = AuditLog.objects.filter(
        timestamp__date=timezone.now().date()
    ).values('user__username').annotate(
        count=Count('id')
    ).filter(user__username__isnull=False).order_by('-count')[:5]
    
    context = {
        'page_obj': page_obj,
        'logs_audit': page_obj.object_list,
        'stats_audit': stats_audit,
        'actions_par_type': actions_par_type,
        'actions_par_utilisateur': actions_par_utilisateur,
        'actions_par_heure': actions_par_heure,
        'actions_par_jour': actions_par_jour,
        'actions_critiques': actions_critiques,
        'utilisateurs_actifs_aujourd_hui': utilisateurs_actifs_aujourd_hui,
        'filters': {
            'search': search_query,
            'action_type': action_type,
            'user': user_filter,
            'date_from': date_from,
            'date_to': date_to,
            'page_size': page_size,
        },
        'total_filtered': queryset.count(),
    }
    
    return render(request, 'core/rapports_audit.html', context)


def export_audit_data(queryset, format_type):
    """
    Export des données d'audit dans différents formats
    """
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    import csv
    import io
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="audit_logs_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Date/Heure', 'Utilisateur', 'Action', 'Type Objet', 'ID Objet', 'Détails', 'IP'])
        
        for log in queryset:
            writer.writerow([
                log.timestamp.strftime('%d/%m/%Y %H:%M:%S'),
                log.user.username if log.user else 'Système',
                log.get_action_display(),
                log.content_type.model if log.content_type else '-',
                log.object_id if log.object_id else '-',
                str(log.details)[:100] if log.details else '-',
                log.ip_address if log.ip_address else '-'
            ])
        
        return response
    
    elif format_type == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = "Logs d'Audit"
        
        # En-têtes
        headers = ['Date/Heure', 'Utilisateur', 'Action', 'Type Objet', 'ID Objet', 'Détails', 'IP']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Données
        for row, log in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=log.timestamp.strftime('%d/%m/%Y %H:%M:%S'))
            ws.cell(row=row, column=2, value=log.user.username if log.user else 'Système')
            ws.cell(row=row, column=3, value=log.get_action_display())
            ws.cell(row=row, column=4, value=log.content_type.model if log.content_type else '-')
            ws.cell(row=row, column=5, value=log.object_id if log.object_id else '-')
            ws.cell(row=row, column=6, value=str(log.details)[:100] if log.details else '-')
            ws.cell(row=row, column=7, value=log.ip_address if log.ip_address else '-')
        
        # Ajustement automatique des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="audit_logs_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
    
    elif format_type == 'pdf':
        # Pour l'export PDF, on peut utiliser reportlab ou weasyprint
        # Pour l'instant, retournons un message d'erreur
        from django.contrib import messages
        messages.error(request, "L'export PDF n'est pas encore disponible.")
        return redirect('core:rapports_audit')
    
    return redirect('core:rapports_audit')


@login_required
def detection_anomalies(request):
    """
    Vue pour détecter et afficher les anomalies du système
    """
    # Vérification des permissions : PRIVILEGE, CONTROLES peuvent accéder à la détection d'anomalies
    from core.utils import check_group_permissions
    permissions = check_group_permissions(request.user, ['PRIVILEGE', 'CONTROLES'], 'view')
    if not permissions['allowed']:
        messages.error(request, permissions['message'])
        return redirect('core:dashboard')
    
    anomalies = []
    
    # 1. Détecter les paiements en retard
    from datetime import date
    from calendar import monthrange
    
    paiements_retard = []
    contrats_actifs = Contrat.objects.filter(est_actif=True).select_related('locataire', 'propriete')
    
    for contrat in contrats_actifs:
        # Calculer la date d'échéance pour le mois en cours
        aujourd_hui = timezone.now().date()
        annee = aujourd_hui.year
        mois = aujourd_hui.month
        
        # Obtenir le dernier jour du mois
        _, dernier_jour_mois = monthrange(annee, mois)
        
        # Date d'échéance = jour de paiement du contrat ou dernier jour du mois si le jour n'existe pas
        jour_paiement = min(contrat.jour_paiement, dernier_jour_mois)
        date_echeance = date(annee, mois, jour_paiement)
        
        # Vérifier si le paiement pour ce mois existe et est en retard
        paiement_mois = Paiement.objects.filter(
            contrat=contrat,
            type_paiement='loyer',
            date_paiement__year=annee,
            date_paiement__month=mois,
            statut='en_attente'
        ).first()
        
        if paiement_mois:
            # Paiement existe mais pas encore reçu
            if aujourd_hui > date_echeance:
                jours_retard = (aujourd_hui - date_echeance).days
                paiements_retard.append(paiement_mois)
                anomalies.append({
                    'type': 'paiement_retard',
                    'severite': 'warning' if jours_retard <= 7 else 'danger',
                    'titre': f'Paiement en retard de {jours_retard} jours',
                    'description': f'Paiement en retard pour {contrat.locataire.nom}',
                    'date': date_echeance,
                    'objet': paiement_mois,
                    'url': f'/paiements/detail/{paiement_mois.id}/'
                })
        else:
            # Aucun paiement enregistré pour ce mois
            if aujourd_hui > date_echeance:
                jours_retard = (aujourd_hui - date_echeance).days
                anomalies.append({
                    'type': 'paiement_manquant',
                    'severite': 'warning' if jours_retard <= 7 else 'danger',
                    'titre': f'Paiement manquant depuis {jours_retard} jours',
                    'description': f'Loyer manquant pour {contrat.locataire.nom}',
                    'date': date_echeance,
                    'objet': contrat,
                    'url': f'/contrats/detail/{contrat.id}/'
                })
    
    # 2. Détecter les contrats expirant bientôt
    date_limite = timezone.now().date() + timedelta(days=30)
    contrats_expirant = Contrat.objects.filter(
        date_fin__lte=date_limite,
        est_actif=True
    ).select_related('locataire', 'propriete')
    
    for contrat in contrats_expirant:
        jours_avant_expiration = (contrat.date_fin - timezone.now().date()).days
        anomalies.append({
            'type': 'contrat_expirant',
            'severite': 'warning' if jours_avant_expiration > 7 else 'danger',
            'titre': f'Contrat expirant dans {jours_avant_expiration} jours',
            'description': f'Contrat {contrat.numero_contrat} pour {contrat.locataire.nom}',
            'date': contrat.date_fin,
            'objet': contrat,
            'url': f'/contrats/detail/{contrat.id}/'
        })
    
    # 3. Détecter les propriétés sans contrat actif
    proprietes_vides = Propriete.objects.exclude(
        contrats__est_actif=True
    ).distinct()
    
    for propriete in proprietes_vides:
        anomalies.append({
            'type': 'propriete_vide',
            'severite': 'info',
            'titre': 'Propriété sans locataire actif',
            'description': f'Propriété {propriete.titre}',
            'date': timezone.now().date(),
            'objet': propriete,
            'url': f'/proprietes/detail/{propriete.id}/'
        })
    
    # 4. Détecter les utilisateurs inactifs
    date_inactivite = timezone.now() - timedelta(days=90)
    utilisateurs_inactifs = Utilisateur.objects.filter(
        last_login__lt=date_inactivite
    ).exclude(is_superuser=True)
    
    for utilisateur in utilisateurs_inactifs:
        jours_inactivite = (timezone.now().date() - utilisateur.last_login.date()).days
        anomalies.append({
            'type': 'utilisateur_inactif',
            'severite': 'warning',
            'titre': f'Utilisateur inactif depuis {jours_inactivite} jours',
            'description': f'Utilisateur {utilisateur.username}',
            'date': utilisateur.last_login,
            'objet': utilisateur,
            'url': f'/utilisateurs/detail/{utilisateur.id}/'
        })
    
    # Trier les anomalies par sévérité et date
    severite_order = {'danger': 0, 'warning': 1, 'info': 2}
    anomalies.sort(key=lambda x: (severite_order.get(x['severite'], 3), x['date']))
    
    # Statistiques des anomalies
    stats_anomalies = {
        'total': len(anomalies),
        'danger': len([a for a in anomalies if a['severite'] == 'danger']),
        'warning': len([a for a in anomalies if a['severite'] == 'warning']),
        'info': len([a for a in anomalies if a['severite'] == 'info']),
    }
    
    context = {
        'anomalies': anomalies,
        'stats_anomalies': stats_anomalies,
        'paiements_retard': paiements_retard,
        'contrats_expirant': contrats_expirant,
        'proprietes_vides': proprietes_vides,
        'utilisateurs_inactifs': utilisateurs_inactifs,
    }
    
    return render(request, 'core/detection_anomalies.html', context)

@login_required
def detail_audit_log(request, log_id):
    """
    Vue pour afficher les détails d'un log d'audit spécifique
    """
    # Vérification des permissions : PRIVILEGE, CONTROLES peuvent accéder aux détails d'audit
    from core.utils import check_group_permissions
    permissions = check_group_permissions(request.user, ['PRIVILEGE', 'CONTROLES'], 'view')
    if not permissions['allowed']:
        messages.error(request, permissions['message'])
        return redirect('core:rapports_audit')
    
    # Récupération du log d'audit
    try:
        log = AuditLog.objects.select_related('user', 'content_type').get(id=log_id)
    except AuditLog.DoesNotExist:
        messages.error(request, "Log d'audit introuvable.")
        return redirect('core:rapports_audit')
    
    # Récupération des informations contextuelles
    context_info = {}
    if log.content_type and log.object_id:
        try:
            # Essayer de récupérer l'objet référencé
            model_class = log.content_type.model_class()
            if model_class:
                obj = model_class.objects.get(id=log.object_id)
                context_info['objet'] = obj
                context_info['objet_url'] = obj.get_absolute_url() if hasattr(obj, 'get_absolute_url') else None
        except:
            context_info['objet_supprime'] = True
    
    # Récupération des logs similaires (même utilisateur, même type d'objet)
    logs_similaires = AuditLog.objects.filter(
        user=log.user,
        content_type=log.content_type,
        object_id=log.object_id
    ).exclude(id=log.id).order_by('-timestamp')[:5]
    
    # Récupération des actions de l'utilisateur récentes
    actions_utilisateur = AuditLog.objects.filter(
        user=log.user
    ).exclude(id=log.id).order_by('-timestamp')[:10]
    
    context = {
        'log': log,
        'context_info': context_info,
        'logs_similaires': logs_similaires,
        'actions_utilisateur': actions_utilisateur,
    }
    
    return render(request, 'core/detail_audit_log.html', context)


@login_required
def audit_statistiques(request):
    """
    Vue pour afficher des statistiques d'audit avancées avec graphiques
    """
    # Vérification des permissions : PRIVILEGE, CONTROLES peuvent accéder aux statistiques d'audit
    from core.utils import check_group_permissions
    permissions = check_group_permissions(request.user, ['PRIVILEGE', 'CONTROLES'], 'view')
    if not permissions['allowed']:
        messages.error(request, permissions['message'])
        return redirect('core:rapports_audit')
    
    # Période de référence
    periode = request.GET.get('periode', '30')  # jours
    date_debut = timezone.now() - timedelta(days=int(periode))
    
    # Statistiques par période
    stats_periode = AuditLog.objects.filter(
        timestamp__gte=date_debut
    ).aggregate(
        total=Count('id'),
        creations=Count('id', filter=Q(action='create')),
        modifications=Count('id', filter=Q(action='update')),
        suppressions=Count('id', filter=Q(action='delete')),
        consultations=Count('id', filter=Q(action='view')),
        connexions=Count('id', filter=Q(action='login')),
    )
    
    # Actions par jour (pour graphique linéaire)
    actions_par_jour = AuditLog.objects.filter(
        timestamp__gte=date_debut
    ).extra(
        select={'date': 'DATE(timestamp)'}
    ).values('date').annotate(
        total=Count('id'),
        creations=Count('id', filter=Q(action='create')),
        modifications=Count('id', filter=Q(action='update')),
        suppressions=Count('id', filter=Q(action='delete')),
    ).order_by('date')
    
    # Actions par heure de la journée
    actions_par_heure = AuditLog.objects.filter(
        timestamp__gte=date_debut
    ).extra(
        select={'heure': 'EXTRACT(hour FROM timestamp)'}
    ).values('heure').annotate(
        count=Count('id')
    ).order_by('heure')
    
    # Top des utilisateurs actifs
    top_utilisateurs = AuditLog.objects.filter(
        timestamp__gte=date_debut
    ).values('user__username').annotate(
        total_actions=Count('id'),
        creations=Count('id', filter=Q(action='create')),
        modifications=Count('id', filter=Q(action='update')),
        suppressions=Count('id', filter=Q(action='delete')),
    ).filter(user__username__isnull=False).order_by('-total_actions')[:10]
    
    # Top des objets les plus consultés/modifiés
    top_objets = AuditLog.objects.filter(
        timestamp__gte=date_debut,
        content_type__isnull=False
    ).values('content_type__model').annotate(
        total_actions=Count('id'),
        consultations=Count('id', filter=Q(action='view')),
        modifications=Count('id', filter=Q(action='update')),
    ).order_by('-total_actions')[:10]
    
    # Anomalies détectées
    anomalies = []
    
    # Utilisateurs avec beaucoup de suppressions
    utilisateurs_suppressions = AuditLog.objects.filter(
        timestamp__gte=date_debut,
        action='delete'
    ).values('user__username').annotate(
        count=Count('id')
    ).filter(count__gte=5)  # Plus de 5 suppressions
    
    for user in utilisateurs_suppressions:
        anomalies.append({
            'type': 'suppressions_multiples',
            'utilisateur': user['user__username'],
            'count': user['count'],
            'description': f"L'utilisateur {user['user__username']} a effectué {user['count']} suppressions"
        })
    
    # Actions en dehors des heures de travail (8h-18h)
    actions_hors_travail = AuditLog.objects.filter(
        timestamp__gte=date_debut
    ).extra(
        where=['EXTRACT(hour FROM timestamp) < 8 OR EXTRACT(hour FROM timestamp) > 18']
    ).count()
    
    if actions_hors_travail > 0:
        anomalies.append({
            'type': 'actions_hors_travail',
            'count': actions_hors_travail,
            'description': f"{actions_hors_travail} actions effectuées en dehors des heures de travail"
        })
    
    context = {
        'periode': periode,
        'date_debut': date_debut,
        'stats_periode': stats_periode,
        'actions_par_jour': list(actions_par_jour),
        'actions_par_heure': list(actions_par_heure),
        'top_utilisateurs': top_utilisateurs,
        'top_objets': top_objets,
        'anomalies': anomalies,
        'actions_hors_travail': actions_hors_travail,
    }
    
    return render(request, 'core/audit_statistiques.html', context)

def test_phone_widget(request):
    """Vue de test pour le composant de sélection de pays et téléphone."""
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    
    # Créer un champ factice pour le composant
    class FakeField:
        def __init__(self):
            self.id_for_label = 'telephone_test'
            self.name = 'telephone'
            self.value = ''
            self.help_text = 'Sélectionnez le pays puis saisissez le numéro'
            self.errors = []
    
    context = {
        'phone_field': FakeField(),
    }
    
    html = render_to_string('test_phone_widget.html', context, request=request)
    return HttpResponse(html)
