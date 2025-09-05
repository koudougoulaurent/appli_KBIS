"""
Vues spécialisées pour la gestion des paiements et retraits avec les unités locatives
"""
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Q, Count, Avg
from django.utils import timezone
from django.core.paginator import Paginator
from datetime import datetime, date, timedelta
from decimal import Decimal

from core.utils import check_group_permissions, get_context_with_entreprise_config
from proprietes.models import Bailleur, UniteLocative
from .models import RetraitBailleur, DetailRetraitUnite, Paiement
from .services_unites_locatives import ServiceUnitesLocativesFinancier
from .forms import RetraitBailleurForm


@login_required
def rapport_unites_locatives(request, bailleur_id=None):
    """
    Affiche un rapport détaillé des unités locatives avec leurs performances financières.
    """
    context = get_context_with_entreprise_config(request)
    
    # Sélection du bailleur
    if bailleur_id:
        bailleur = get_object_or_404(Bailleur, pk=bailleur_id, is_deleted=False)
    else:
        # Si aucun bailleur spécifié, prendre le premier disponible
        bailleur = Bailleur.objects.filter(is_deleted=False).first()
        if not bailleur:
            messages.error(request, "Aucun bailleur disponible.")
            return redirect('paiements:dashboard')
    
    # Paramètres de période
    mois_actuel = timezone.now().date().replace(day=1)
    mois_param = request.GET.get('mois')
    if mois_param:
        try:
            mois_analyse = datetime.strptime(mois_param, '%Y-%m').date().replace(day=1)
        except ValueError:
            mois_analyse = mois_actuel
    else:
        mois_analyse = mois_actuel
    
    # Calcul des revenus par unité
    revenus_par_unite = ServiceUnitesLocativesFinancier.calculer_revenus_par_unite(
        bailleur, mois_analyse
    )
    
    # Statistiques globales
    stats_globales = {
        'nombre_unites_total': revenus_par_unite['totaux']['nombre_unites_total'],
        'nombre_unites_occupees': revenus_par_unite['totaux']['nombre_unites_occupees'],
        'taux_occupation': revenus_par_unite['totaux'].get('taux_occupation', 0),
        'revenus_totaux': revenus_par_unite['totaux']['revenus_nets'],
        'loyers_theoriques': revenus_par_unite['totaux']['loyers_bruts'],
        'charges_deductibles': revenus_par_unite['totaux']['charges_deductibles']
    }
    
    # Calcul du taux de recouvrement global
    if stats_globales['loyers_theoriques'] > 0:
        stats_globales['taux_recouvrement'] = (
            stats_globales['revenus_totaux'] / stats_globales['loyers_theoriques'] * 100
        )
    else:
        stats_globales['taux_recouvrement'] = 0
    
    # Classement des unités par performance
    unites_classees = sorted(
        revenus_par_unite['unites'],
        key=lambda x: x['revenus_nets'],
        reverse=True
    )
    
    # Identification des unités problématiques
    unites_problematiques = [
        unite for unite in revenus_par_unite['unites']
        if unite['statut_paiement'] in ['Partiel', 'Impayé']
    ]
    
    # Liste des bailleurs pour le sélecteur
    tous_bailleurs = Bailleur.objects.filter(is_deleted=False).order_by('nom')
    
    # Mois disponibles pour navigation
    mois_precedent = mois_analyse - timedelta(days=1)
    mois_precedent = mois_precedent.replace(day=1)
    
    if mois_analyse.month == 12:
        mois_suivant = date(mois_analyse.year + 1, 1, 1)
    else:
        mois_suivant = date(mois_analyse.year, mois_analyse.month + 1, 1)
    
    context.update({
        'title': f'Rapport des Unités Locatives - {bailleur.nom}',
        'bailleur': bailleur,
        'mois_analyse': mois_analyse,
        'mois_precedent': mois_precedent,
        'mois_suivant': mois_suivant,
        'tous_bailleurs': tous_bailleurs,
        'stats_globales': stats_globales,
        'unites_classees': unites_classees,
        'unites_problematiques': unites_problematiques,
        'revenus_par_unite': revenus_par_unite
    })
    
    return render(request, 'paiements/rapport_unites_locatives.html', context)


@login_required
def creer_retrait_avec_unites(request, bailleur_id):
    """
    Crée un retrait en détaillant les contributions de chaque unité locative.
    """
    bailleur = get_object_or_404(Bailleur, pk=bailleur_id, is_deleted=False)
    context = get_context_with_entreprise_config(request)
    
    # Mois pour le retrait
    mois_retrait = timezone.now().date().replace(day=1)
    mois_param = request.GET.get('mois')
    if mois_param:
        try:
            mois_retrait = datetime.strptime(mois_param, '%Y-%m').date().replace(day=1)
        except ValueError:
            pass
    
    if request.method == 'POST':
        form = RetraitBailleurForm(request.POST)
        if form.is_valid():
            # Calcul détaillé avec unités
            calculs_retrait = ServiceUnitesLocativesFinancier.calculer_retrait_avec_unites(
                bailleur, mois_retrait
            )
            
            # Créer le retrait principal
            retrait = form.save(commit=False)
            retrait.bailleur = bailleur
            retrait.mois_retrait = mois_retrait
            retrait.cree_par = request.user
            retrait.montant_loyers_bruts = calculs_retrait['calculs_retrait']['revenus_bruts_total']
            retrait.montant_charges_deductibles = calculs_retrait['charges_bailleur']
            retrait.montant_net_a_payer = calculs_retrait['calculs_retrait']['montant_net_retrait']
            retrait.save()
            
            # Créer les détails par unité
            for unite_data in calculs_retrait['revenus_par_unite']['unites']:
                DetailRetraitUnite.objects.create(
                    retrait=retrait,
                    unite_locative=unite_data['unite'],
                    contrat=unite_data['contrat'],
                    loyer_theorique=unite_data['loyer_theorique'],
                    charges_theoriques=unite_data['charges_theoriques'],
                    paiements_recus=unite_data['paiements_recus'],
                    charges_deductibles=unite_data['charges_deductibles'],
                    revenus_nets_unite=unite_data['revenus_nets'],
                    taux_paiement=unite_data['taux_paiement'],
                    statut_paiement=unite_data['statut_paiement'].lower()
                )
            
            messages.success(
                request, 
                f'Retrait créé avec succès pour {bailleur.nom}. '
                f'Montant net: {retrait.montant_net_a_payer} F CFA'
            )
            return redirect('paiements:detail_retrait_avec_unites', retrait_id=retrait.pk)
    else:
        # Pré-calcul pour affichage
        calculs_retrait = ServiceUnitesLocativesFinancier.calculer_retrait_avec_unites(
            bailleur, mois_retrait
        )
        
        initial_data = {
            'montant_loyers_bruts': calculs_retrait['calculs_retrait']['revenus_bruts_total'],
            'montant_charges_deductibles': calculs_retrait['charges_bailleur'],
            'montant_net_a_payer': calculs_retrait['calculs_retrait']['montant_net_retrait'],
            'type_retrait': 'mensuel',
            'statut': 'en_attente'
        }
        form = RetraitBailleurForm(initial=initial_data)
    
    context.update({
        'title': f'Nouveau Retrait - {bailleur.nom}',
        'bailleur': bailleur,
        'form': form,
        'mois_retrait': mois_retrait,
        'calculs_retrait': calculs_retrait if 'calculs_retrait' in locals() else None
    })
    
    return render(request, 'paiements/creer_retrait_avec_unites.html', context)


@login_required
def detail_retrait_avec_unites(request, retrait_id):
    """
    Affiche le détail d'un retrait avec la répartition par unité locative.
    """
    retrait = get_object_or_404(RetraitBailleur, pk=retrait_id, is_deleted=False)
    context = get_context_with_entreprise_config(request)
    
    # Récupérer les détails par unité
    details_unites = retrait.details_unites.all().select_related(
        'unite_locative', 'contrat', 'contrat__locataire'
    )
    
    # Calculs statistiques
    stats_retrait = {
        'nombre_unites': details_unites.count(),
        'unites_rentables': details_unites.filter(revenus_nets_unite__gt=0).count(),
        'revenus_moyen_par_unite': details_unites.aggregate(
            avg=Avg('revenus_nets_unite')
        )['avg'] or Decimal('0'),
        'taux_paiement_moyen': details_unites.aggregate(
            avg=Avg('taux_paiement')
        )['avg'] or Decimal('0'),
        'total_loyers_theoriques': details_unites.aggregate(
            total=Sum('loyer_theorique')
        )['total'] or Decimal('0'),
        'total_paiements_recus': details_unites.aggregate(
            total=Sum('paiements_recus')
        )['total'] or Decimal('0')
    }
    
    # Répartition par statut de paiement
    repartition_statuts = {}
    for statut_code, statut_label in DetailRetraitUnite._meta.get_field('statut_paiement').choices:
        count = details_unites.filter(statut_paiement=statut_code).count()
        repartition_statuts[statut_label] = count
    
    # Unités les plus et moins rentables
    unite_plus_rentable = details_unites.order_by('-revenus_nets_unite').first()
    unite_moins_rentable = details_unites.order_by('revenus_nets_unite').first()
    
    context.update({
        'title': f'Détail du Retrait - {retrait.bailleur.nom}',
        'retrait': retrait,
        'details_unites': details_unites,
        'stats_retrait': stats_retrait,
        'repartition_statuts': repartition_statuts,
        'unite_plus_rentable': unite_plus_rentable,
        'unite_moins_rentable': unite_moins_rentable
    })
    
    return render(request, 'paiements/detail_retrait_avec_unites.html', context)


@login_required
def api_donnees_unites_locatives(request, bailleur_id):
    """
    API pour récupérer les données des unités locatives en JSON.
    Utilisé pour les graphiques et tableaux dynamiques.
    """
    bailleur = get_object_or_404(Bailleur, pk=bailleur_id, is_deleted=False)
    
    # Paramètres
    mois = request.GET.get('mois')
    if mois:
        try:
            mois_analyse = datetime.strptime(mois, '%Y-%m').date().replace(day=1)
        except ValueError:
            mois_analyse = timezone.now().date().replace(day=1)
    else:
        mois_analyse = timezone.now().date().replace(day=1)
    
    # Calcul des données
    revenus_par_unite = ServiceUnitesLocativesFinancier.calculer_revenus_par_unite(
        bailleur, mois_analyse
    )
    
    # Formatage pour JSON
    donnees = {
        'mois': mois_analyse.strftime('%Y-%m'),
        'bailleur': {
            'id': bailleur.id,
            'nom': bailleur.nom
        },
        'totaux': {
            'nombre_unites_total': revenus_par_unite['totaux']['nombre_unites_total'],
            'nombre_unites_occupees': revenus_par_unite['totaux']['nombre_unites_occupees'],
            'taux_occupation': float(revenus_par_unite['totaux'].get('taux_occupation', 0)),
            'revenus_totaux': float(revenus_par_unite['totaux']['revenus_nets']),
            'loyers_bruts': float(revenus_par_unite['totaux']['loyers_bruts']),
            'charges_deductibles': float(revenus_par_unite['totaux']['charges_deductibles'])
        },
        'unites': []
    }
    
    for unite_data in revenus_par_unite['unites']:
        donnees['unites'].append({
            'id': unite_data['unite'].id,
            'numero': unite_data['unite'].numero_unite,
            'nom': unite_data['unite'].nom,
            'propriete': unite_data['unite'].propriete.titre,
            'locataire': unite_data['locataire'].nom if unite_data['locataire'] else None,
            'loyer_theorique': float(unite_data['loyer_theorique']),
            'charges_theoriques': float(unite_data['charges_theoriques']),
            'paiements_recus': float(unite_data['paiements_recus']),
            'charges_deductibles': float(unite_data['charges_deductibles']),
            'revenus_nets': float(unite_data['revenus_nets']),
            'taux_paiement': float(unite_data['taux_paiement']),
            'statut_paiement': unite_data['statut_paiement']
        })
    
    return JsonResponse(donnees)


@login_required
def export_rapport_unites_excel(request, bailleur_id):
    """
    Exporte le rapport des unités locatives au format Excel.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    
    bailleur = get_object_or_404(Bailleur, pk=bailleur_id, is_deleted=False)
    
    # Paramètres
    mois = request.GET.get('mois')
    if mois:
        try:
            mois_analyse = datetime.strptime(mois, '%Y-%m').date().replace(day=1)
        except ValueError:
            mois_analyse = timezone.now().date().replace(day=1)
    else:
        mois_analyse = timezone.now().date().replace(day=1)
    
    # Calcul des données
    revenus_par_unite = ServiceUnitesLocativesFinancier.calculer_revenus_par_unite(
        bailleur, mois_analyse
    )
    
    # Création du workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Unités {mois_analyse.strftime('%Y-%m')}"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center")
    
    # En-têtes
    headers = [
        'Unité', 'Propriété', 'Locataire', 'Loyer Théorique', 'Charges Théoriques',
        'Paiements Reçus', 'Charges Déductibles', 'Revenus Nets', 'Taux Paiement (%)',
        'Statut'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Données
    for row, unite_data in enumerate(revenus_par_unite['unites'], 2):
        ws.cell(row=row, column=1, value=unite_data['unite'].numero_unite)
        ws.cell(row=row, column=2, value=unite_data['unite'].propriete.titre)
        ws.cell(row=row, column=3, value=unite_data['locataire'].nom if unite_data['locataire'] else 'Vacant')
        ws.cell(row=row, column=4, value=float(unite_data['loyer_theorique']))
        ws.cell(row=row, column=5, value=float(unite_data['charges_theoriques']))
        ws.cell(row=row, column=6, value=float(unite_data['paiements_recus']))
        ws.cell(row=row, column=7, value=float(unite_data['charges_deductibles']))
        ws.cell(row=row, column=8, value=float(unite_data['revenus_nets']))
        ws.cell(row=row, column=9, value=float(unite_data['taux_paiement']))
        ws.cell(row=row, column=10, value=unite_data['statut_paiement'])
    
    # Ajustement des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="unites_locatives_{bailleur.nom}_{mois_analyse.strftime("%Y_%m")}.xlsx"'
    )
    
    wb.save(response)
    return response
